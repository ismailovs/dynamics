from __future__ import annotations

import tempfile
import unittest
from dataclasses import replace
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from urllib.error import HTTPError

from openpyxl import load_workbook

from youtube_trend_pipeline.config import Settings
from youtube_trend_pipeline.database import Database
from youtube_trend_pipeline.export import export_workbook
from youtube_trend_pipeline.filtering import filter_videos
from youtube_trend_pipeline.models import Video
from youtube_trend_pipeline.pipeline import TrendPipeline
from youtube_trend_pipeline.youtube import (
    QuotaBudgetExceeded,
    YouTubeClient,
    parse_duration,
)

NOW = datetime(2026, 7, 16, 10, 0, tzinfo=timezone.utc)


def make_video(
    index: int,
    topic: str = "ai",
    **overrides: object,
) -> Video:
    if topic == "ai":
        title = f"How AI Agents Transform Factory Work Episode {index}"
        description = (
            "Inside the future of artificial intelligence agents and automation "
            "for office and factory work."
        )
        category = "AI and technology"
    else:
        title = f"Why Nuclear Energy Is Returning Project {index}"
        description = (
            "The engineering of nuclear reactors and the future of clean energy "
            "infrastructure."
        )
        category = "Energy and infrastructure"
    values: dict[str, object] = {
        "video_id": f"{topic}-{index}",
        "title": title,
        "description": description,
        "channel_id": f"channel-{topic}-{index % 3}",
        "channel_title": f"Channel {topic} {index % 3}",
        "published_at": NOW - timedelta(days=(index % 5) + 1),
        "duration_seconds": 900 + index,
        "view_count": 10_000 + index * 1_000,
        "like_count": 500 + index * 10,
        "comment_count": 50 + index,
        "subscriber_count": 20_000 + index * 100,
        "tags": [topic, "documentary", "engineering"],
        "language": "en",
        "thumbnail_url": f"https://example.com/{topic}-{index}.jpg",
        "category": "28",
        "discovery_category": category,
    }
    values.update(overrides)
    return Video(**values)  # type: ignore[arg-type]


def _api_video_item(
    video_id: str, duration: str = "PT12M"
) -> dict[str, object]:
    return {
        "id": video_id,
        "snippet": {
            "title": f"How Technology Changes the World {video_id}",
            "description": "Inside the engineering and technology story.",
            "channelId": f"channel-{video_id}",
            "channelTitle": "Channel",
            "publishedAt": "2026-07-10T00:00:00Z",
            "categoryId": "28",
            "defaultLanguage": "en",
        },
        "contentDetails": {"duration": duration},
        "statistics": {
            "viewCount": "1000",
            "likeCount": "50",
            "commentCount": "5",
        },
    }


class DurationAndFilterTests(unittest.TestCase):
    def test_iso_duration_parser(self) -> None:
        self.assertEqual(parse_duration("PT8M"), 480)
        self.assertEqual(parse_duration("PT1H2M3S"), 3723)

    def test_filters_unsuitable_and_duplicate_videos(self) -> None:
        good = make_video(1)
        duplicate = make_video(2, title=good.title, channel_id=good.channel_id)
        short = make_video(3, duration_seconds=120)
        stream = make_video(4, live_broadcast_content="live")
        music = make_video(5, category="10")
        non_english = make_video(6, language="es")
        missing_stats = make_video(7, like_count=-1)

        result = filter_videos(
            [good, duplicate, short, stream, music, non_english, missing_stats]
        )

        self.assertEqual([video.video_id for video in result.accepted], ["ai-1"])
        self.assertEqual(result.rejected["duplicate_channel_title"], 1)
        self.assertEqual(result.rejected["duration"], 1)
        self.assertEqual(result.rejected["livestream"], 1)
        self.assertEqual(result.rejected["music"], 1)
        self.assertEqual(result.rejected["non_english"], 1)
        self.assertEqual(result.rejected["missing_statistics"], 1)


class YouTubeClientTests(unittest.TestCase):
    def test_default_budget_reserves_quota_for_detail_requests(self) -> None:
        self.assertEqual(Settings().max_search_requests, 90)

    def test_discovery_sends_rolling_window_and_deduplicates(self) -> None:
        calls: list[tuple[str, dict[str, object]]] = []

        def request(resource: str, params: dict[str, object]) -> dict[str, object]:
            calls.append((resource, params))
            return {
                "items": [
                    {"id": {"videoId": "one"}},
                    {"id": {"videoId": "one"}},
                ]
            }

        client = YouTubeClient("test", request_json=request)
        result = client.discover(
            NOW - timedelta(days=14),
            NOW,
            max_search_requests=2,
        )

        self.assertEqual(result, {"one": "AI and technology"})
        self.assertEqual(len(calls), 2)
        self.assertEqual(calls[0][1]["maxResults"], 50)
        self.assertEqual(calls[0][1]["order"], "date")
        self.assertEqual(calls[0][1]["type"], "video")
        self.assertEqual(calls[0][1]["publishedAfter"], "2026-07-02T10:00:00Z")
        self.assertEqual(calls[0][1]["publishedBefore"], "2026-07-16T10:00:00Z")

    def test_discovery_balances_first_pages_before_pagination(self) -> None:
        calls: list[dict[str, object]] = []

        def request(_resource: str, params: dict[str, object]) -> dict[str, object]:
            calls.append(params)
            return {"items": [], "nextPageToken": f"next-{params['q']}"}

        client = YouTubeClient("test", request_json=request)
        client.discover(
            NOW - timedelta(days=14),
            NOW,
            max_search_requests=61,
        )

        self.assertEqual(len(calls), 61)
        self.assertTrue(all("pageToken" not in call for call in calls[:60]))
        self.assertIn("pageToken", calls[60])

    def test_hidden_subscriber_count_is_preserved_as_unknown(self) -> None:
        def request(resource: str, _params: dict[str, object]) -> dict[str, object]:
            if resource == "channels":
                return {
                    "items": [
                        {
                            "id": "channel",
                            "statistics": {"hiddenSubscriberCount": True},
                        }
                    ]
                }
            return {
                "items": [
                    {
                        "id": "video",
                        "snippet": {
                            "title": "How Technology Is Changing the World",
                            "description": "Inside the engineering story.",
                            "channelId": "channel",
                            "channelTitle": "Channel",
                            "publishedAt": "2026-07-10T00:00:00Z",
                            "categoryId": "28",
                        },
                        "contentDetails": {"duration": "PT12M"},
                        "statistics": {
                            "viewCount": "1000",
                            "likeCount": "50",
                            "commentCount": "5",
                        },
                    }
                ]
            }

        video = YouTubeClient("test", request_json=request).fetch_videos(
            {"video": "AI and technology"}
        )[0]
        self.assertEqual(video.subscriber_count, -1)

    def test_retries_429_with_exponential_backoff(self) -> None:
        attempts = 0
        delays: list[float] = []

        def request(_resource: str, _params: dict[str, object]) -> dict[str, object]:
            nonlocal attempts
            attempts += 1
            if attempts < 4:
                raise HTTPError("https://example.test", 429, "rate limited", {}, None)
            return {"items": []}

        client = YouTubeClient("test", request_json=request, sleep=delays.append)
        client.search_page("robots", NOW - timedelta(days=1), NOW)

        self.assertEqual(attempts, 4)
        self.assertEqual(delays, [1.0, 2.0, 4.0])

    def test_video_details_are_batched_at_fifty_ids(self) -> None:
        video_batch_sizes: list[int] = []

        def request(resource: str, params: dict[str, object]) -> dict[str, object]:
            if resource == "channels":
                return {"items": []}
            ids = str(params["id"]).split(",")
            video_batch_sizes.append(len(ids))
            return {"items": [_api_video_item(video_id) for video_id in ids]}

        discovered = {f"video-{index}": "AI and technology" for index in range(51)}
        videos = YouTubeClient("test", request_json=request).fetch_videos(discovered)

        self.assertEqual(len(videos), 51)
        self.assertEqual(video_batch_sizes, [50, 1])


class DiscoveryCacheTests(unittest.TestCase):
    def test_failed_attempts_consume_persisted_quota_and_stop_retries(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            attempts = 0
            delays: list[float] = []

            def request(
                _resource: str, _params: dict[str, object]
            ) -> dict[str, object]:
                nonlocal attempts
                attempts += 1
                raise HTTPError("https://example.test", 429, "rate limited", {}, None)

            with Database(Path(directory) / "trends.db") as database:
                client = YouTubeClient(
                    "test",
                    request_json=request,
                    sleep=delays.append,
                    usage_recorder=lambda resource: database.reserve_api_request(
                        resource, NOW.date(), 200
                    ),
                )
                with self.assertRaises(QuotaBudgetExceeded):
                    client.search_page("robots", NOW - timedelta(days=1), NOW)

                self.assertEqual(attempts, 2)
                self.assertEqual(delays, [1.0, 2.0])
                self.assertEqual(database.quota_units_used(NOW.date()), 200)

    def test_pending_details_are_deferred_when_quota_is_reserved(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            settings = replace(
                Settings(),
                database_path=Path(directory) / "trends.db",
                daily_quota_units=1,
            )
            calls = 0

            def request(
                _resource: str, _params: dict[str, object]
            ) -> dict[str, object]:
                nonlocal calls
                calls += 1
                return {"items": []}

            with Database(settings.database_path) as database:
                database.record_search_page(
                    "AI and technology",
                    "robots",
                    NOW - timedelta(days=14),
                    NOW,
                    ["one", "two"],
                    "",
                    discovered_at=NOW,
                )
                result = TrendPipeline(settings, database).collect(
                    YouTubeClient("test", request_json=request), now=NOW
                )

                self.assertEqual(result.accepted, [])
                self.assertEqual(calls, 0)
                self.assertEqual(len(database.pending_discoveries()), 2)

    def test_refresh_defers_batches_after_quota_is_consumed(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            settings = replace(
                Settings(),
                database_path=Path(directory) / "trends.db",
                daily_quota_units=1,
            )
            batch_sizes: list[int] = []

            def request(resource: str, params: dict[str, object]) -> dict[str, object]:
                self.assertEqual(resource, "videos")
                ids = str(params["id"]).split(",")
                batch_sizes.append(len(ids))
                return {
                    "items": [
                        {
                            "id": video_id,
                            "statistics": {
                                "viewCount": "2000",
                                "likeCount": "100",
                                "commentCount": "10",
                            },
                        }
                        for video_id in ids
                    ]
                }

            videos = [make_video(index) for index in range(51)]
            with Database(settings.database_path) as database:
                pipeline = TrendPipeline(settings, database)
                pipeline.ingest(videos, collected_at=NOW)
                saved = pipeline.refresh(
                    YouTubeClient("test", request_json=request), now=NOW
                )

                self.assertEqual(saved, 50)
                self.assertEqual(batch_sizes, [50])
                self.assertEqual(database.quota_units_used(NOW.date()), 1)

    def test_daily_quota_survives_restart_and_oldest_queries_advance(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            settings = replace(
                Settings(),
                database_path=Path(directory) / "trends.db",
                daily_quota_units=206,
            )
            searched_queries: list[str] = []

            def request(resource: str, params: dict[str, object]) -> dict[str, object]:
                self.assertEqual(resource, "search")
                searched_queries.append(str(params["q"]))
                return {"items": []}

            with Database(settings.database_path) as database:
                pipeline = TrendPipeline(settings, database)
                pipeline.collect(
                    YouTubeClient("test", request_json=request), now=NOW
                )
                first_day_queries = searched_queries.copy()
                self.assertEqual(len(first_day_queries), 2)
                self.assertEqual(database.quota_units_used(NOW.date()), 200)

                pipeline.collect(
                    YouTubeClient("test", request_json=request), now=NOW
                )
                self.assertEqual(searched_queries, first_day_queries)

                next_day = NOW + timedelta(days=1)
                pipeline.collect(
                    YouTubeClient("test", request_json=request), now=next_day
                )
                second_day_queries = searched_queries[2:]
                self.assertEqual(len(second_day_queries), 2)
                self.assertTrue(
                    set(first_day_queries).isdisjoint(second_day_queries)
                )

    def test_collection_saves_all_ids_and_reuses_completed_search_cache(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            settings = replace(
                Settings(),
                database_path=Path(directory) / "trends.db",
            )
            calls: list[tuple[str, dict[str, object]]] = []

            def request(resource: str, params: dict[str, object]) -> dict[str, object]:
                calls.append((resource, params))
                if resource == "search":
                    return {
                        "items": [
                            {"id": {"videoId": "accepted"}},
                            {"id": {"videoId": "accepted"}},
                            {"id": {"videoId": "too-short"}},
                        ]
                    }
                if resource == "channels":
                    return {"items": []}
                return {
                    "items": [
                        _api_video_item("accepted"),
                        _api_video_item("too-short", duration="PT2M"),
                    ]
                }

            client = YouTubeClient("test", request_json=request)
            with Database(settings.database_path) as database:
                pipeline = TrendPipeline(settings, database)
                first = pipeline.collect(client, now=NOW)
                search_calls = sum(resource == "search" for resource, _ in calls)
                detail_calls = sum(resource == "videos" for resource, _ in calls)

                self.assertEqual(len(first.accepted), 1)
                self.assertEqual(len(database.discovered_videos()), 2)
                statuses = {
                    row["video_id"]: row["details_status"]
                    for row in database.discovered_videos()
                }
                self.assertEqual(
                    statuses, {"accepted": "accepted", "too-short": "rejected"}
                )
                self.assertEqual(search_calls, 60)
                self.assertEqual(detail_calls, 1)

                second = pipeline.collect(client, now=NOW)

                self.assertEqual(second.accepted, [])
                self.assertEqual(
                    sum(resource == "search" for resource, _ in calls),
                    search_calls,
                )
                self.assertEqual(
                    sum(resource == "videos" for resource, _ in calls),
                    detail_calls,
                )

    def test_incomplete_search_resumes_from_cached_page_token(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            with Database(Path(directory) / "trends.db") as database:
                database.record_search_page(
                    "AI and technology",
                    "robots",
                    NOW - timedelta(days=14),
                    NOW,
                    ["one"],
                    "next-page",
                    discovered_at=NOW,
                )
                work = database.search_work(
                    "AI and technology",
                    "robots",
                    NOW - timedelta(days=14),
                    NOW,
                )

                self.assertIsNotNone(work)
                self.assertEqual(work[2], "next-page")  # type: ignore[index]


class EndToEndTests(unittest.TestCase):
    def test_database_analysis_and_six_sheet_export(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            settings = replace(
                Settings(),
                database_path=root / "nested" / "trends.db",
                report_path=root / "trends.xlsx",
                target_clusters=2,
                min_theme_videos=5,
            )
            videos = [
                make_video(index, "ai", subscriber_count=-1) for index in range(5)
            ]
            videos += [make_video(index, "nuclear") for index in range(5)]
            historical = [
                make_video(
                    index,
                    "nuclear",
                    video_id=f"historical-{index}",
                    title=f"The Forgotten History Archive Chapter {index}",
                    description="The history of an ancient archive and its lost records.",
                    published_at=NOW - timedelta(days=30 + index),
                    discovery_category="History and mysteries",
                )
                for index in range(5)
            ]

            with Database(settings.database_path) as database:
                pipeline = TrendPipeline(settings, database)
                result = pipeline.ingest(
                    videos + historical, collected_at=NOW - timedelta(days=1)
                )
                self.assertEqual(len(result.accepted), 15)
                refreshed_ids: list[str] = []

                class RefreshClient:
                    def refresh_statistics(
                        self, video_ids: object
                    ) -> dict[str, tuple[int, int, int]]:
                        refreshed_ids.extend(video_ids)  # type: ignore[arg-type]
                        return {
                            video.video_id: (
                                video.view_count,
                                video.like_count,
                                video.comment_count,
                            )
                            for video in videos
                        }

                self.assertEqual(
                    pipeline.refresh(RefreshClient(), now=NOW),  # type: ignore[arg-type]
                    10,
                )
                self.assertEqual(set(refreshed_ids), {video.video_id for video in videos})
                self.assertEqual(pipeline._search_request_budget(50_000), 87)
                database.snapshot(snapshot_date=date(2026, 7, 15))
                increased = {
                    video.video_id: (
                        video.view_count + 2_000 + index * 100,
                        video.like_count + 20,
                        video.comment_count + 2,
                    )
                    for index, video in enumerate(videos)
                }
                database.snapshot(increased, snapshot_date=date(2026, 7, 16))
                partial_video = videos[0]
                self.assertEqual(
                    database.snapshot(
                        {
                            partial_video.video_id: (
                                partial_video.view_count + 3_000,
                                -1,
                                -1,
                            ),
                            "unknown-video": (100, 5, 1),
                        },
                        snapshot_date=date(2026, 7, 16),
                    ),
                    1,
                )

                self.assertEqual(pipeline.analyze(now=NOW), 2)
                themes = database.themes()
                self.assertTrue(all(theme["median_view_velocity"] > 0 for theme in themes))
                self.assertTrue(all(theme["video_count"] == 5 for theme in themes))
                ai_theme = next(
                    theme
                    for theme in themes
                    if theme["parent_category"] == "AI and technology"
                )
                self.assertIsNone(ai_theme["median_view_subscriber_ratio"])
                partial_row = next(
                    row
                    for row in database.videos()
                    if row["video_id"] == partial_video.video_id
                )
                self.assertEqual(partial_row["like_count"], partial_video.like_count + 20)
                report = export_workbook(database, settings.report_path)

            workbook = load_workbook(report, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Top 1000 Themes",
                    "Best Video Examples",
                    "Fastest Growing",
                    "Underserved Themes",
                    "Video Data",
                    "Methodology",
                ],
            )
            self.assertEqual(workbook["Video Data"].max_row, 16)
            self.assertEqual(workbook["Top 1000 Themes"].max_row, 3)


if __name__ == "__main__":
    unittest.main()
